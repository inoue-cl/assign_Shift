import os
from flask import (
    Flask,
    request,
    redirect,
    url_for,
    render_template_string,
    send_file,
)
import gspread
from google.oauth2.service_account import Credentials
import csv
from io import BytesIO, TextIOWrapper
from openpyxl import Workbook

app = Flask(__name__)

SCOPES = ["https://www.googleapis.com/auth/spreadsheets"]


def get_client():
    creds_file = os.environ.get("GOOGLE_SERVICE_ACCOUNT_FILE")
    if not creds_file:
        raise RuntimeError("GOOGLE_SERVICE_ACCOUNT_FILE not set")
    credentials = Credentials.from_service_account_file(creds_file, scopes=SCOPES)
    return gspread.authorize(credentials)


def get_sheets():
    sheet_id = os.environ.get("SPREADSHEET_ID")
    if not sheet_id:
        raise RuntimeError("SPREADSHEET_ID not set")
    client = get_client()
    sh = client.open_by_key(sheet_id)
    people = sh.worksheet("People")
    projects = sh.worksheet("Projects")
    assignments = sh.worksheet("Assignments")
    return people, projects, assignments


def load_people(sheet):
    return [row[0] for row in sheet.get_all_values()[1:]]


def load_projects(sheet):
    return [row[0] for row in sheet.get_all_values()[1:]]


def load_assignments(sheet):
    rows = sheet.get_all_values()[1:]
    result = []
    for p, proj, month, frac in rows:
        result.append({"person": p, "project": proj, "month": month, "fraction": frac})
    return result


@app.route("/")
def index():
    people_s, projects_s, assignments_s = get_sheets()
    assignments = load_assignments(assignments_s)
    return render_template_string(
        """
        <h1>Assignments</h1>
        <table border=1>
        <tr><th>Person</th><th>Project</th><th>Month</th><th>Fraction</th></tr>
        {% for a in assignments %}
        <tr><td>{{a.person}}</td><td>{{a.project}}</td><td>{{a.month}}</td><td>{{a.fraction}}</td></tr>
        {% endfor %}
        </table>
        <p><a href='{{ url_for("add_person") }}'>Add Person</a></p>
        <p><a href='{{ url_for("add_project") }}'>Add Project</a></p>
        <p><a href='{{ url_for("add_assignment") }}'>Add Assignment</a></p>
        """,
        assignments=assignments,
    )


@app.route("/people", methods=["GET", "POST"])
def add_person():
    people_s, _, _ = get_sheets()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            people_s.append_row([name])
        return redirect(url_for("index"))
    people = load_people(people_s)
    return render_template_string(
        """
        <h1>People</h1>
        <ul>{% for p in people %}<li>{{p}}</li>{% endfor %}</ul>
        <form method='post'>
        <input name='name' placeholder='Name'>
        <input type='submit' value='Add'>
        </form>
        <p><a href='{{ url_for("index") }}'>Back</a></p>
        """,
        people=people,
    )


@app.route("/projects", methods=["GET", "POST"])
def add_project():
    _, projects_s, _ = get_sheets()
    if request.method == "POST":
        name = request.form.get("name", "").strip()
        if name:
            projects_s.append_row([name])
        return redirect(url_for("index"))
    projects = load_projects(projects_s)
    return render_template_string(
        """
        <h1>Projects</h1>
        <ul>{% for p in projects %}<li>{{p}}</li>{% endfor %}</ul>
        <form method='post'>
        <input name='name' placeholder='Name'>
        <input type='submit' value='Add'>
        </form>
        <p><a href='{{ url_for("index") }}'>Back</a></p>
        """,
        projects=projects,
    )


@app.route("/assign", methods=["GET", "POST"])
def add_assignment():
    people_s, projects_s, assignments_s = get_sheets()
    if request.method == "POST":
        person = request.form.get("person")
        project = request.form.get("project")
        month = request.form.get("month")
        fraction = request.form.get("fraction")
        if person and project and month and fraction:
            assignments_s.append_row([person, project, month, fraction])
        return redirect(url_for("index"))
    people = load_people(people_s)
    projects = load_projects(projects_s)
    return render_template_string(
        """
        <h1>Add Assignment</h1>
        <form method='post'>
        <label>Person</label>
        <select name='person'>{% for p in people %}<option>{{p}}</option>{% endfor %}</select><br>
        <label>Project</label>
        <select name='project'>{% for p in projects %}<option>{{p}}</option>{% endfor %}</select><br>
        <label>Month (YYYY-MM)</label>
        <input name='month'><br>
        <label>Fraction</label>
        <input name='fraction' type='number' step='0.1' min='0' max='1'><br>
        <input type='submit' value='Add'>
        </form>
        <p><a href='{{ url_for("index") }}'>Back</a></p>
        """,
        people=people,
        projects=projects,
    )


@app.route("/export_excel")
def export_excel():
    """Export all data to an Excel file with three sheets."""
    people_s, projects_s, assignments_s = get_sheets()
    wb = Workbook()

    ws_people = wb.active
    ws_people.title = "People"
    ws_people.append(["Name"])
    for p in load_people(people_s):
        ws_people.append([p])

    ws_proj = wb.create_sheet("Projects")
    ws_proj.append(["Name"])
    for p in load_projects(projects_s):
        ws_proj.append([p])

    ws_assign = wb.create_sheet("Assignments")
    ws_assign.append(["Person", "Project", "Month", "Fraction"])
    for a in load_assignments(assignments_s):
        ws_assign.append([a["person"], a["project"], a["month"], a["fraction"]])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        download_name="assignments.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/import_csv", methods=["GET", "POST"])
def import_csv():
    """Import assignment data from a CSV file uploaded by the user."""
    people_s, projects_s, assignments_s = get_sheets()
    if request.method == "POST" and "file" in request.files:
        file = request.files["file"]
        if file.filename:
            wrapper = TextIOWrapper(file.stream, encoding="utf-8")
            reader = csv.DictReader(wrapper)
            for row in reader:
                person = row.get("Person")
                project = row.get("Project")
                month = row.get("Month")
                fraction = row.get("Fraction")
                if person and project and month and fraction:
                    assignments_s.append_row([person, project, month, fraction])
        return redirect(url_for("index"))
    return render_template_string(
        """
        <h1>Import CSV</h1>
        <form method='post' enctype='multipart/form-data'>
        <input type='file' name='file' accept='.csv'>
        <input type='submit' value='Upload'>
        </form>
        <p><a href='{{ url_for("index") }}'>Back</a></p>
        """
    )


if __name__ == "__main__":
    app.run(debug=True)
